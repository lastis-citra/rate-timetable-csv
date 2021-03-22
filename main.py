import csv
import os
from datetime import date
from datetime import datetime
import re
import logging

import cloudscraper
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side


logger = logging.getLogger(__name__)


# HTMLをダウンロードしてsoupを返す
def download_html(url, file_path):
    scraper = cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'desktop': True
        }
    )
    res = scraper.get(url)
    # print(html)

    # TODO: 保存されるHTMLがSJISになっている
    with open(file_path, mode='w') as f:
        f.write(res.text)

    soup = BeautifulSoup(res.content, 'html.parser')
    return soup


# ダウンロード済みのHTMLを読み込んでsoupを返す
def open_cache(path):
    soup = BeautifulSoup(open(path), 'html.parser')
    return soup


# 結果をExcelに出力する
# https://qiita.com/orengepy/items/d10ad53fee5593b29e46
def output_excel(result_list, types_list, excel_path, color_setting):
    wb = Workbook()
    ws = wb.active

    # 最も長い行数を求める
    max_x = 0
    for results in result_list:
        if max_x < len(results):
            max_x = len(results)

    # 最も長い行に合わせて空白を足す
    results_list_added = []
    for results in result_list:
        lack = max_x - len(results)
        results_added = results
        for _ in range(lack):
            results_added.append("")
        results_list_added.append(results_added)

    types_list_added = []
    for types in types_list:
        lack = max_x - len(types)
        types_added = types
        for _ in range(lack):
            types_added.append("")
        types_list_added.append(types_added)

    dest_font = Font(name='メイリオ', size=8, color='000000')
    min_font = Font(name='メイリオ', size=11, color='000000')
    dest_align = Alignment(horizontal='center', vertical='top')
    min_align = Alignment(horizontal='center', vertical='center')
    odd_fill = PatternFill(patternType='solid', fgColor='ffffff')
    even_fill = PatternFill(patternType='solid', fgColor='f2f2f2')
    side = Side(style='thin', color='000000')
    dest_border = Border(top=side)
    min_border = Border(bottom=side)

    def set_color(_y, _x, _types_list):
        with open(color_setting, 'r', errors='replace', encoding="utf_8") as file:
            d = dict(filter(None, csv.reader(file)))

        train_type = _types_list[_y][_x]
        type_color = d[train_type]
        return Font(name=min_font.name, size=min_font.size, color=type_color)

    def write_list_2d(sheet, list_2d, start_row, start_col):
        for y, row in enumerate(list_2d):
            for x, cell in enumerate(row):
                row = start_row + y
                col = start_col + x
                sheet.cell(row=row, column=col, value=list_2d[y][x])

                # 行き先行のフォント設定
                if y % 2 == 0:
                    sheet.cell(row=row, column=col).font = dest_font
                    sheet.cell(row=row, column=col).alignment = dest_align
                    sheet.cell(row=row, column=col).border = dest_border
                # 時刻行のフォント設定
                else:
                    sheet.cell(row=row, column=col).font = set_color(y // 2, x, types_list_added)
                    sheet.cell(row=row, column=col).alignment = min_align
                    sheet.cell(row=row, column=col).border = min_border
                # 白い行
                if y % 4 == 0 or y % 4 == 1:
                    sheet.cell(row=row, column=col).fill = odd_fill
                # 灰色の行
                else:
                    sheet.cell(row=row, column=col).fill = even_fill

    write_list_2d(ws, results_list_added, 2, 3)

    wb.save(excel_path)
    wb.close()


def create_time_table(table_soup, excel_path, dest_setting, color_setting):
    # 行き先を省略して1文字にする
    def replace_dests(_dests):
        with open(dest_setting, 'r', errors='replace', encoding="utf_8") as file:
            d = dict(filter(None, csv.reader(file)))

        _dests_replaced = []
        for _dest in _dests:
            # 辞書dのkeyに一致するものがあれば，そのvalueで置き換える
            if _dest in d:
                _dests_replaced.append(d[_dest])
            else:
                _dests_replaced.append(_dest)

        return _dests_replaced

    # 当駅始発の場合は●に置換する
    def replace_starts(_starts):
        starts_replaced = []
        for _start in _starts:
            if _start == '':
                starts_replaced.append(_start)
            else:
                starts_replaced.append("●")

        return starts_replaced

    tds = table_soup.select('tr.ek-hour_line td')
    # tdのlistから偶数番目を取り出して，それぞれのtextを取り出している
    # 時刻のリスト，特に使っていない
    hours = list(map(lambda x: x.text, tds[0::2]))
    print(hours)

    # 種別のリスト
    types_list = []
    # 行き先のリスト
    dests_list = []
    # 分のリスト
    mins_list = []
    # tdのlistから奇数番目を取り出して処理する
    for trains in tds[1::2]:
        types = list(map(lambda x: x['data-tr-type'], trains.select('li.ek-tooltip')))
        types_list.append(types)

        dests = list(map(lambda x: x['data-dest'], trains.select('li.ek-tooltip')))
        starts = list(map(lambda x: x['data-start'], trains.select('li.ek-tooltip')))

        dests_replaced = []
        for dest, start in zip(replace_dests(dests), replace_starts(starts)):
            dests_replaced.append(start + dest)

        dests_list.append(dests_replaced)

        mins = list(map(lambda x: x.text, trains.select('span.time-min')))
        mins_list.append(mins)

    # print(dests_list)
    # print(types_list)
    # print(mins_list)

    result_list = []
    for dests, mins in zip(dests_list, mins_list):
        print(','.join(dests))
        print(','.join(mins))
        result_list.append(dests)
        result_list.append(mins)

    output_excel(result_list, types_list, excel_path, color_setting)


def prepare_soup(url, html_dir, excel_dir, name, dw, dest_setting, color_setting):
    today = date.today()
    today_string = today.strftime('%Y%m%d')

    html_name = today_string + '_' + name + '_' + dw + '.html'
    html_path = os.path.join(html_dir, html_name)

    # 当日のキャッシュがある場合はキャッシュを利用し，なければダウンロードする
    if os.path.exists(html_path):
        soup = open_cache(html_path)
    else:
        soup = download_html(url, html_path)

    # 時刻表の更新日時を取得
    updated_date_text = soup.select_one('div.date time').text
    updated_date_tuple = re.search(r'(\d+)年(\d+)月(\d+)日現在', updated_date_text).groups()
    u_year = updated_date_tuple[0]
    u_mon = updated_date_tuple[1]
    u_day = updated_date_tuple[2]
    updated_date = datetime(int(u_year), int(u_mon), int(u_day)).strftime('%Y%m%d')
    # print(updated_date)

    # TODO: 全部同じExcelにしてシートだけ分けたい
    excel_name = updated_date + '_' + name + '_' + dw
    excel_path_up = os.path.join(excel_dir, excel_name + '_up.xlsx')
    excel_path_down = os.path.join(excel_dir, excel_name + '_down.xlsx')

    # <tr class="ek-hour_line">
    #   <td>07</td>
    #   <td>
    #     <ul>
    #       <li class="ek-tooltip ek-narrow ek-train-tooltip" data-tr-type="普通" data-dest="宇都宮" data-kind_palette="t1" data-start="" data-link01="#">
    #         <a href="#" data-sf="2709" data-tx="570110-16593-2520Y" data-departure="0706" class="tooltip-data ek-train-link t1" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">
    #           <span class="dest means-text" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">[普通]宇</span>
    #           <span class="time-min means-text" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">06</span>
    #         </a>
    #       </li>
    #       <li class="ek-tooltip ek-narrow ek-train-tooltip" data-tr-type="普通" data-dest="宇都宮" data-kind_palette="t1" data-start="" data-link01="#">
    #         <a href="#" data-sf="2709" data-tx="570110-16596-2522Y" data-departure="0750" class="tooltip-data ek-train-link t1" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">
    #           <span class="dest means-text" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">[普通]宇</span>
    #           <span class="time-min means-text" ga-event-lbl="GA-TRAC_railway-line-station-pocket_PC_result-time">50</span>
    #         </a>
    #       </li>
    #     </ul>
    #   </td>
    # </tr>
    tables = soup.select('div.search-result-body')

    # 上り
    if not os.path.exists(excel_path_up):
        create_time_table(tables[0], excel_path_up, dest_setting, color_setting)

    # 下り
    if not os.path.exists(excel_path_down):
        create_time_table(tables[1], excel_path_down, dest_setting, color_setting)


def main_function(file_name, html_dir, excel_dir, setting_dir):
    with open(file_name, 'r', errors='replace', encoding="utf_8") as file:
        line_list = file.readlines()

    line_count = 0

    for line in line_list:
        line_count += 1
        input_url = line.split(',')[0]
        file_name = line.split(',')[1]
        dest_setting = os.path.join(setting_dir, line.split(',')[2])
        color_setting = os.path.join(setting_dir, line.split(',')[3].replace('\n', ''))
        print(line_count, '/', len(line_list))
        print('input_url: ' + input_url)

        # 平日分
        input_url1 = input_url + '?dw=0'
        prepare_soup(input_url1, html_dir, excel_dir, file_name, 'weekday', dest_setting, color_setting)
        # 休日分
        input_url2 = input_url + '?dw=2'
        prepare_soup(input_url2, html_dir, excel_dir, file_name, 'holiday', dest_setting, color_setting)


if __name__ == '__main__':
    html_directory = 'html'
    excel_directory = 'excel'
    setting_directory = 'setting'
    input_file_name = './input_url_list.txt'

    if not os.path.exists(html_directory):
        os.makedirs(html_directory)

    if not os.path.exists(excel_directory):
        os.makedirs(excel_directory)

    if not os.path.exists(setting_directory):
        logging.error('There is no setting directory!!!')
        exit(1)

    main_function(input_file_name, html_directory, excel_directory, setting_directory)
